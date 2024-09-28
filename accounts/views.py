from django.shortcuts import render,redirect
from .models import *
from django.contrib import messages
from django.contrib import auth
from django.contrib.auth.decorators import login_required
from django.contrib.sites.shortcuts import get_current_site
from django.template.loader import render_to_string
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes
from django.contrib.auth.tokens import default_token_generator
from django.core.mail import EmailMessage
from django.contrib.auth.forms import SetPasswordForm
from django.contrib.auth import update_session_auth_hash
from .filters import NameFilter
from django.db import transaction

import openpyxl

from django.core.validators import validate_email
from django.core.exceptions import ValidationError
import random



def main(request):
    return render(request,'accounts/main.html')

@login_required(login_url='all_login')
def add_emp(request):
    return render(request,'add_emp.html')


@login_required(login_url='all_login')
@transaction.atomic
def process_excel_file(request, sheet):
    fname_col = 1
    lname_col = 2
    email_col = 3

    accounts_to_create = []

    for row in range(2, sheet.max_row + 1):
        fname = sheet.cell(row=row, column=fname_col).value
        lname = sheet.cell(row=row, column=lname_col).value
        email = sheet.cell(row=row, column=email_col).value

        try:
            validate_email(email)
            account, created = Account.objects.get_or_create(
                email=email,
                defaults={
                    'first_name': fname,
                    'last_name': lname,
                    'company_name': request.user.company_name
                }
            )

            if created:
                # Set password, save, and send email only if the account is newly created
                password = "123456"
                account.set_password(password)
                account.is_employee= True
                account.is_active=True
                account.is_staff=True
                account.save()

                current_site = get_current_site(request)
                mail_subject = "You have been Registered"
                message = render_to_string('accounts/emp_email.html', {
                    'user': account,
                    'domain': current_site.domain,
                    'uid': urlsafe_base64_encode(force_bytes(account.pk)),
                    'token': default_token_generator.make_token(account)
                })

                to_email = account.email
                send_email = EmailMessage(mail_subject, message, to=[to_email])
                send_email.send()
            else:
                messages.warning(request, f'{email} already exists and was not added.')

        except ValidationError:
            messages.warning(request, f'{email} not added due to invalid format')

    messages.success(request, f' Employees created')
    return redirect('add_emp')


def bulk_register(request):
    if request.method == 'POST':
        excel = request.FILES.get('excel')
        if excel:
            try:
                workbook = openpyxl.load_workbook(excel)
                sheet = workbook.active
                return process_excel_file(request, sheet)
            except Exception as e:
                messages.error(request, f'Error processing Excel file: {str(e)}')
    else:
        messages.error(request, 'No Excel file provided.')

    return render(request, 'add_emp.html')

@login_required(login_url='all_login')
def single_reg(request):
    if request.method=='POST':
        fname = request.POST['fname']
        lname = request.POST['lname']
        email = request.POST['email']
        if Account.objects.filter(email=email).exists():
            messages.error(request,f'User with email {email} already exists')
        else:
            emp = Account.objects.create_user(first_name=fname,last_name=lname,email=email)
            emp.company_name = request.user.company_name
            password="123456"
            emp.set_password(password)
            emp.save()
            current_site=get_current_site(request)
            mail_subject="You have been Registered"
            message=render_to_string('accounts/emp_email.html',
                        {'user':emp,
                        'domain':current_site,
                        'uid':urlsafe_base64_encode(force_bytes(emp.pk)),
                        'token':default_token_generator.make_token(emp)})
            to_email=email
            send_email=EmailMessage(mail_subject,message,to=[to_email])
            send_email.send()
        return redirect('add_emp')


@login_required(login_url='all_login')
def emp_list(request):
    records = Account.objects.filter(company_name=request.user.company_name,is_admin=False)
    my_filter = NameFilter(request.GET,queryset=records)
    records=my_filter.qs
    context = {
        'records':records,
        'my_filter':my_filter,
    }
    return render(request,'emp_list.html',context)


@login_required(login_url='all_login')
def  emp_delete(request,pk):
    if request.user.is_authenticated:
        record = Account.objects.get(id=pk)
        record.delete()
        messages.success(request,f"All the To-do's assigned to {record.first_name} Deleted Successfully!!!")
        return redirect('emp_list')
    else:
        messages.error(request,'You must be logged in to delete a record ')
        return redirect('home')
    



def all_login(request):
    if request.method == 'POST':
        email = request.POST['email']
        password = request.POST['password']
        user = auth.authenticate(request, email=email, password=password)
        if user is not None:
            auth.login(request, user)
            return redirect('home')
        elif user is None:
            messages.error(request, 'Invalid Email or password.')

    return render(request,'accounts/company_login.html')

def company_register(request):
    if request.method =='POST':
        fname=request.POST['fname']
        lname=request.POST['lname']
        designation=request.POST['designation']
        phone_number=request.POST['phone_number']
        email=request.POST['email']
        # password=request.POST['password']
        # confirm_passowrd=request.POST['c_password']
        company_name=request.POST['company_name']

        val_num = ['7','8','9']
        if Account.objects.filter(email=email).exists():
            messages.error(request,f'User with email {email} already exists')
        elif Account.objects.filter(company_name=company_name).exists():
            messages.error(request,f'Company {company_name} already exists')
        elif len(phone_number)>10:
            messages.error(request,'Invalid, Phone number must be of 10 digits')
        elif len(phone_number)<10:
            messages.error(request,'Invalid, Phone number must be of 10 digits')
        elif phone_number[0] not in val_num:
            messages.error(request,'Invalid, Phone Number should stat with 7,8 or 9')
        else:
            company = Account.objects.create_superuser(first_name=fname,last_name=lname,phone_number=phone_number,designation = designation,email=email,company_name=company_name)
            password="123456"
            company.set_password(password)
            company.save()
            

            
            messages.success(request,f"Please check your email {email} for Login Id and Password")
            current_site=get_current_site(request)
            mail_subject="Please activate your account"
            message=render_to_string('accounts/account_verification_email.html',
                        {'user':company,
                        'domain':current_site,
                        'uid':urlsafe_base64_encode(force_bytes(company.pk)),
                        'token':default_token_generator.make_token(company)})
            to_email=email
            send_email=EmailMessage(mail_subject,message,to=[to_email])
            send_email.send()

    return render(request,'accounts/company_reg.html')




def activate(request,uid,token):#ye code hai decryption ke liye
    try:
       uid=urlsafe_base64_decode(uid).decode()
       user=Account._default_manager.get(pk=uid)
    except Exception:
        user=None
    if user is not None and default_token_generator.check_token(user,token):
        user.is_active=True
        user.save()
        messages.success(request,"Congratulations,Your account is activated")
        return redirect('all_login')
    else:
        messages.error(request,"Oops... Session timed out,Please register again.")
        return redirect('company_register')

def forgotpassword(request):
    if request.method=='POST':
        email=request.POST['email']
        if Account.objects.filter(email=email).exists():
            user=Account.objects.get(email__exact=email)
            current_site=get_current_site(request)
            mail_subject="Reset your password"
            message=render_to_string('accounts/reset_password_email.html',
                         {'user':user,
                          'domain':current_site,
                          'uid':urlsafe_base64_encode(force_bytes(user.pk)),
                          'token':default_token_generator.make_token(user)})
            to_email=email
            send_email=EmailMessage(mail_subject,message,to=[to_email])
            send_email.send()
            messages.success(request,f'Password reset email has been sent to {email}')
            return redirect('all_login')
        else:
            messages.error(request,'Account does not exist')
            return redirect('forgotpassword')
    return render(request,'accounts/forgotpassword.html')


def resetpassword_validate(request,uid,token):
    try:
       uid=urlsafe_base64_decode(uid).decode()
       user=Account._default_manager.get(pk=uid)
    except Exception:
        user=None
    if user is not None and default_token_generator.check_token(user,token):
        request.session['uid']=uid
        user.save()
        messages.success(request,"Please reset your password")
        return redirect('resetpassword')
    else:
        messages.error(request,"link expired")
        return redirect('company_register')
    
def resetpassword(request):
    if request.method=='POST':
        password=request.POST['password']
        confirm_password=request.POST['confirm_password']
        if password==confirm_password:
            uid=request.session.get('uid')
            user=Account.objects.get(pk=uid)
            user.set_password(password)
            user.save()
            messages.success(request,'Password reset was successful')
            return redirect('all_login')
        else:
            messages.error(request,'Password do not match')
            return redirect('resetpassword')
    return render(request,'accounts/resetpassword.html')

def changepassword(request):
    if request.user.is_authenticated:
        if request.method=='POST':
            form = SetPasswordForm(user=request.user,data=request.POST)
            if form.is_valid():
                form.save()
                update_session_auth_hash(request,form.user)
                messages.success(request,'Password Changed')
                return redirect('home')
        else:
            form=SetPasswordForm(user=request.user)
        context = {
            'form':form,
        }
        return render(request,'accounts/changepassword.html',context)
    else:
        messages.error(request,'You must be authenticated')



@login_required(login_url='company_login')
def all_logout(request):
    auth.logout(request)
    return redirect('all_login')
    